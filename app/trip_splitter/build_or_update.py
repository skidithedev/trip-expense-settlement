"""
build_or_update.py
------------------
Build an Excel workbook with trip settlement results.

Steps:
1. Load CSVs
2. Run pipeline (FX conversion → allocations → balances → settlement)
3. Write results to Trip_Splitter.xlsx
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from io import BytesIO

from trip_splitter.logic import (
    load_all_data,
    convert_expenses_to_base,
    compute_allocations,
    compute_balances,
    compute_settlement,
)
from trip_splitter.schemas import TRIP_NAME


OUTPUT_FILE = "Trip_Splitter.xlsx"


def write_df_to_sheet(ws, df, title: str = None, bold_header: bool = True, freeze: bool = True):
    """
    Write a pandas DataFrame to an openpyxl worksheet.
    """
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        start_row = 3
    else:
        start_row = 1

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        ws.append(row)
        if bold_header and r_idx == start_row:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True)

    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def format_currency_column(ws, col_idx: int, start_row: int = 4):
    """
    Apply VND currency formatting to a column in a worksheet.
    """
    VND_FORMAT = '#,##0" ₫"'
    for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = VND_FORMAT

def format_expenses_amount(ws, df, amount_col: str = "Amount", currency_col: str = "Currency", start_row: int = 4):
    """
    Apply per-row currency formatting for the 'Amount' column in Expenses sheet.

    Parameters
    ----------
    ws : openpyxl worksheet
    df : pandas DataFrame (expenses with Currency column)
    amount_col : str, default 'Amount'
        Name of the original amount column.
    currency_col : str, default 'Currency'
        Name of the currency column.
    start_row : int, default=4
        First row of data (after title and header).
    """
    CURRENCY_FORMATS = {
        "VND": '#,##0" ₫"',
        "CNY": '#,##0" ¥"',
        "USD": '"$"#,##0',
        "EUR": '#,##0" €"',
    }

    amount_idx = list(df.columns).index(amount_col) + 1
    currency_idx = list(df.columns).index(currency_col) + 1

    for row_offset, row in enumerate(ws.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1), start=0):
        amount_cell = row[amount_idx - 1]
        currency_cell = row[currency_idx - 1]

        cur = currency_cell.value
        if isinstance(amount_cell.value, (int, float)) and cur in CURRENCY_FORMATS:
            amount_cell.number_format = CURRENCY_FORMATS[cur]

def build_workbook(data_dir="sample_data", out_path=OUTPUT_FILE):
    """
    Run pipeline and build the Excel workbook.
    """
    # Load raw CSVs
    data = load_all_data(data_dir)

    # Pipeline
    expenses_vnd = convert_expenses_to_base(data["expenses"], data["rates"])
    allocations = compute_allocations(expenses_vnd, data["splits"], data["participants"])
    balances = compute_balances(expenses_vnd, allocations, data["participants"])
    settlement = compute_settlement(balances)

    # Create workbook
    wb = Workbook()

    # Sheets in desired order
    ws_settle = wb.active
    ws_settle.title = "Settlement"
    ws_summary = wb.create_sheet("Summary")
    ws_balances = wb.create_sheet("Balances")
    ws_allocs = wb.create_sheet("Allocations")
    ws_expenses = wb.create_sheet("Expenses")

    # Write data
    write_df_to_sheet(ws_settle, settlement, title=f"{TRIP_NAME} Settlement")
    write_df_to_sheet(ws_balances, balances, title="Participant Balances")
    write_df_to_sheet(ws_allocs, allocations, title="Allocations (per expense)")
    write_df_to_sheet(ws_expenses, expenses_vnd, title="Expenses with VND Conversion")

    # Add receipt hyperlinks
    url_col = list(expenses_vnd.columns).index("DriveURL") + 1
    for row_idx, url in enumerate(expenses_vnd["DriveURL"], start=4):
        if pd.notna(url) and str(url).strip():
            cell = ws_expenses.cell(row=row_idx, column=url_col)
            cell.value = "🧾 receipt"
            cell.hyperlink = url
            cell.style = "Hyperlink"

    # --- Summary content ---
    totals_by_cat = expenses_vnd.groupby("Category")["Amount_Base"].sum().reset_index()
    totals_by_person = balances[["Participant", "Paid_Base", "Owed_Base", "Net_Base"]]

    write_df_to_sheet(ws_summary, totals_by_cat, title="Totals by Category")
    start_row = ws_summary.max_row + 2
    for r_idx, row in enumerate(dataframe_to_rows(totals_by_person, index=False, header=True), start=start_row):
        ws_summary.append(row)
        if r_idx == start_row:
            for cell in ws_summary[r_idx]:
                cell.font = Font(bold=True)

    # Charts
    pie = PieChart()
    labels = Reference(ws_summary, min_col=1, min_row=4, max_row=3 + len(totals_by_cat))
    data = Reference(ws_summary, min_col=2, min_row=3, max_row=3 + len(totals_by_cat))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Spending by Category"
    ws_summary.add_chart(pie, "D4")

    bar = BarChart()
    labels = Reference(ws_summary, min_col=1, min_row=start_row + 1, max_row=start_row + len(totals_by_person))
    data = Reference(ws_summary, min_col=2, min_row=start_row, max_col=4, max_row=start_row + len(totals_by_person))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = "Paid vs Owed vs Net"
    ws_summary.add_chart(bar, "D15")

    # Reorder sheets
    wb._sheets = [ws_settle, ws_summary, ws_balances, ws_allocs, ws_expenses]

    # Apply formatting
    amt_col = list(settlement.columns).index("Amount_VND") + 1
    format_currency_column(ws_settle, amt_col)
    for col in ["Paid_Base", "Owed_Base", "Net_Base"]:
        col_idx = list(balances.columns).index(col) + 1
        format_currency_column(ws_balances, col_idx)
    col_idx = list(allocations.columns).index("Share_Base") + 1
    format_currency_column(ws_allocs, col_idx)
    col_idx = list(expenses_vnd.columns).index("Amount_Base") + 1
    format_currency_column(ws_expenses, col_idx)
    
    # Original Amount column formatting with proper currency symbols
    format_expenses_amount(ws_expenses, expenses_vnd, amount_col="Amount", currency_col="Currency")

    # Save
    wb.save(out_path)
    print(f"Workbook saved to {os.path.abspath(out_path)}")


if __name__ == "__main__":
    build_workbook()


def build_workbook_bytes(data_dir="sample_data") -> bytes:
    """
    Build the workbook and return it as bytes for GUI download.
    """
    # Load raw CSVs
    data = load_all_data(data_dir)

    # Pipeline
    expenses_vnd = convert_expenses_to_base(data["expenses"], data["rates"])
    allocations = compute_allocations(expenses_vnd, data["splits"], data["participants"])
    balances = compute_balances(expenses_vnd, allocations, data["participants"])
    settlement = compute_settlement(balances)

    # Create workbook
    wb = Workbook()

    # Sheets in desired order
    ws_settle = wb.active
    ws_settle.title = "Settlement"
    ws_summary = wb.create_sheet("Summary")
    ws_balances = wb.create_sheet("Balances")
    ws_allocs = wb.create_sheet("Allocations")
    ws_expenses = wb.create_sheet("Expenses")

    # Write data
    write_df_to_sheet(ws_settle, settlement, title=f"{TRIP_NAME} Settlement")
    write_df_to_sheet(ws_balances, balances, title="Participant Balances")
    write_df_to_sheet(ws_allocs, allocations, title="Allocations (per expense)")
    write_df_to_sheet(ws_expenses, expenses_vnd, title="Expenses with VND Conversion")

    # Add receipt hyperlinks
    url_col = list(expenses_vnd.columns).index("DriveURL") + 1
    for row_idx, url in enumerate(expenses_vnd["DriveURL"], start=4):
        if pd.notna(url) and str(url).strip():
            cell = ws_expenses.cell(row=row_idx, column=url_col)
            cell.value = "🧾 receipt"
            cell.hyperlink = url
            cell.style = "Hyperlink"

    # --- Summary content ---
    totals_by_cat = expenses_vnd.groupby("Category")["Amount_Base"].sum().reset_index()
    totals_by_person = balances[["Participant", "Paid_Base", "Owed_Base", "Net_Base"]]

    write_df_to_sheet(ws_summary, totals_by_cat, title="Totals by Category")
    start_row = ws_summary.max_row + 2
    for r_idx, row in enumerate(dataframe_to_rows(totals_by_person, index=False, header=True), start=start_row):
        ws_summary.append(row)
        if r_idx == start_row:
            for cell in ws_summary[r_idx]:
                cell.font = Font(bold=True)

    # Charts
    pie = PieChart()
    labels = Reference(ws_summary, min_col=1, min_row=4, max_row=3 + len(totals_by_cat))
    chart_data = Reference(ws_summary, min_col=2, min_row=3, max_row=3 + len(totals_by_cat))
    pie.add_data(chart_data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Spending by Category"
    ws_summary.add_chart(pie, "D4")

    bar = BarChart()
    labels = Reference(ws_summary, min_col=1, min_row=start_row + 1, max_row=start_row + len(totals_by_person))
    chart_data = Reference(ws_summary, min_col=2, min_row=start_row, max_col=4, max_row=start_row + len(totals_by_person))
    bar.add_data(chart_data, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = "Paid vs Owed vs Net"
    ws_summary.add_chart(bar, "D15")

    # Reorder sheets
    wb._sheets = [ws_settle, ws_summary, ws_balances, ws_allocs, ws_expenses]

    # Apply formatting
    amt_col = list(settlement.columns).index("Amount_VND") + 1
    format_currency_column(ws_settle, amt_col)
    for col in ["Paid_Base", "Owed_Base", "Net_Base"]:
        col_idx = list(balances.columns).index(col) + 1
        format_currency_column(ws_balances, col_idx)
    col_idx = list(allocations.columns).index("Share_Base") + 1
    format_currency_column(ws_allocs, col_idx)
    col_idx = list(expenses_vnd.columns).index("Amount_Base") + 1
    format_currency_column(ws_expenses, col_idx)

    # Original Amount column formatting with proper currency symbols
    format_expenses_amount(ws_expenses, expenses_vnd, amount_col="Amount", currency_col="Currency")

    # Save to bytes
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()